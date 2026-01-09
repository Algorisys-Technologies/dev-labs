"""
AI-Powered Financial Data Extractor using OpenAI
Extracts financial data from HTML/Markdown tables using GPT models
"""
import os
import logging
import json
from pathlib import Path
from typing import Dict, List, Optional
import openai
from openai import OpenAI
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

_log = logging.getLogger(__name__)


class AIFinancialExtractor:
    """Extract financial data from parsed documents using OpenAI."""
    
    # System prompt for financial data extraction (template-aware)
    SYSTEM_PROMPT = """You are an expert financial data extraction AI. Extract COMPLETE financial data from quarterly/annual financial statements.

**OUTPUT FORMAT:**
```json
{{
  "company_name": "COMPANY_NAME",
  "financial_data": [
    {{
      "particular": "Sale of goods",
      "key": "sale_of_goods",
      "values": {{
        "30_06_2025_q": "4,357.64",
        "30_06_2024_q": "3,892.15",
        "31_03_2025_q": "15,678.90",
        "31_03_2025_y": "16,859.22"
      }}
    }}
  ]
}}
```

**METRIC CATEGORIES & KEYS:**

Revenue (Operating):
• sale_of_goods, export_sales, service_revenue, other_operating_revenues
• revenue_from_operations = SUM(above) - EXCLUDING other_income

Revenue (Non-Operating):
• other_income - SEPARATE from operations (interest, dividends, gains)
• total_income = revenue_from_operations + other_income

Expenses:
• cost_of_materials_consumed, excise_duty, purchases_stock_in_trade, changes_in_inventories
• employee_benefits_expense, finance_costs, depreciation_amortisation_expense
• other_expense, advertising_expense, impairment_losses, total_expenses

Profit & Tax:
• profit_before_exceptional_and_tax, exceptional_item_expense, profit_before_tax
• current_tax, deferred_tax, total_tax_expense, net_profit

Other:
• other_comprehensive_income, total_comprehensive_income
• paid_up_equity_share_capital, other_equity, eps_basic, eps_diluted

**PERIOD KEY FORMAT - CRITICAL:**
Convert period headers to normalized keys (lowercase, underscores, no special chars).

Examples of normalization:
- "30.06.2025 Q" → "30_06_2025_q"
- "Quarter Ended June 30, 2025" → "quarter_ended_june_30_2025"
- "Year Ended March 31, 2025" → "year_ended_march_31_2025"
- "3M 30th June 2025" → "3m_30th_june_2025"
- "FY 2025" → "fy_2025"

✗ WRONG (using original header):
  "Quarter Ended (Unaudited)": "1,42,064"

✓ CORRECT (normalized key):
  "quarter_ended_unaudited": "1,42,064"

**HOW TO NORMALIZE:**
1. Convert to lowercase
2. Replace spaces/dots with underscores
3. Remove special characters
4. Remove extra underscores

**OTHER EXTRACTION RULES:**
1. Numbers: Keep commas - "4,357.64" not "4357.64"
2. Negatives: Brackets "(123.45)" or minus "-123.45"
3. Missing Values: Use empty string ""
4. Extract ALL rows and ALL periods - no skipping

**CRITICAL:** Return ONLY valid JSON. No explanatory text."""

    # System prompt WITH template metrics (used when template is provided)
    SYSTEM_PROMPT_WITH_TEMPLATE = """You are an expert financial data extraction AI. Extract COMPLETE financial data from quarterly/annual financial statements.

**CRITICAL: Extract ALL date columns/periods found in the table - NOT just 2 periods!**

**STEP 1: IDENTIFY ALL DATE COLUMNS**
First, scan the table header and identify EVERY date column (e.g., 30.06.2025, 30.06.2024, 31.03.2025, 31.12.2024).
Tables typically have 3-4+ periods. Extract values for ALL of them.

**STEP 2: EXTRACT FOR REQUIRED METRICS**
You will receive a list of required metrics from the Excel template.
For EACH metric, extract values for ALL periods identified in Step 1.

REQUIRED METRICS:
{template_metrics}

**OUTPUT FORMAT:**
```json
{{
  "company_name": "COMPANY_NAME",
  "financial_data": [
    {{
      "particular": "Sale of goods",
      "key": "sale_of_goods",
      "values": {{
        "30_06_2025_q": "4,357.64",
        "30_06_2024_q": "3,892.15",
        "31_03_2025_q": "15,678.90",
        "31_03_2025_y": "16,859.22"
      }}
    }}
  ]
}}
```

**MATCHING STRATEGY:**
✓ GOOD MATCHES (accept these):
  "Cost of material consumed" ≈ "Cost of materials consumed"
  "Employee benefit expense" ≈ "Employee benefits expense"  
  "PBT" ≈ "Profit before tax"

✓ MATCH CRITERIA (priority order):
  1. Exact match (case-insensitive)
  2. Partial word match
  3. Synonym match (PBT = Profit Before Tax)
  4. Abbreviation match (singular/plural)
  5. Semantic similarity (>75%)

**PERIOD KEY FORMAT - CRITICAL:**
You will receive REQUIRED_PERIODS mapping that shows:
- formatted_key: The key to use in JSON output
- original_header: The text to look for in the table

Example mapping:
{{
  "30_06_2025_q": "30.06.2025 Q",
  "quarter_ended_june_30_2025": "Quarter Ended June 30, 2025",
  "3m_30th_june_2025": "3M 30th June 2025"
}}

✗ WRONG (using original header as key):
  {{"Quarter Ended June 30, 2025": "1,42,064"}}

✓ CORRECT (using formatted_key from mapping):
  {{"quarter_ended_june_30_2025": "1,42,064"}}

**HOW TO EXTRACT:**
1. Look at REQUIRED_PERIODS mapping
2. Find the original_header text in the table column
3. Extract the value from that column
4. Use the formatted_key (not original_header) in JSON output

**OTHER EXTRACTION RULES:**
1. Numbers: Keep commas - "4,357.64" not "4357.64"
2. Negatives: Brackets "(123.45)" or minus "-123.45"
3. Missing Values: Use empty string ""
4. Extract ALL periods found in table header - minimum 3-4 periods
5. If metric in template but not in table → include with empty values {{}}

**CRITICAL:** Return ONLY valid JSON. No explanatory text."""

    USER_PROMPT_TEMPLATE = """Extract ALL financial data from this {format} table for {company_name}:

{content}

Return ONLY the JSON object. No additional text."""

    USER_PROMPT_WITH_TEMPLATE = """Extract financial data from this {format} table for {company_name}.

**REQUIRED_PERIODS (use formatted_key in JSON output):**
{period_mapping}

**INSTRUCTIONS:**
1. For each metric (from REQUIRED METRICS in system prompt)
2. Find the metric row in the table
3. For each period in REQUIRED_PERIODS:
   - Find the column with the original_header text
   - Extract the value from that cell
   - Use the formatted_key (not original_header) in your JSON output
4. Return complete data for ALL metrics and ALL periods

TABLE CONTENT:
{content}

Return ONLY the JSON object. Use formatted_key for all period keys in "values" object. No explanations."""

    def __init__(self, api_key: Optional[str] = None, model: str = "gpt-4o-mini"):
        """
        Initialize AI extractor.
        
        Args:
            api_key: OpenAI API key (defaults to OPENAI_API_KEY env var)
            model: OpenAI model to use (default: gpt-4o-mini for cost efficiency)
        """
        self.api_key = api_key or os.getenv('OPENAI_API_KEY')
        if not self.api_key:
            raise ValueError("OpenAI API key not provided. Set OPENAI_API_KEY environment variable.")
        
        self.client = OpenAI(api_key=self.api_key)
        self.model = model
        _log.info(f"Initialized AIFinancialExtractor with model: {model}")
    
    def _read_file_content(self, file_path: Path) -> str:
        """Read content from HTML or Markdown file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            _log.error(f"Error reading file {file_path}: {str(e)}")
            raise
    
    def _clean_json_response(self, response: str) -> str:
        """Clean OpenAI response to extract valid JSON."""
        # Remove markdown code blocks if present
        response = response.strip()
        if response.startswith('```json'):
            response = response[7:]
        elif response.startswith('```'):
            response = response[3:]
        if response.endswith('```'):
            response = response[:-3]
        return response.strip()
    
    def _is_section_heading(self, cell_value: str, cell) -> bool:
        """
        Detect if a cell is a section heading, not a metric.
        
        Uses multiple signals:
        1. Bold formatting (primary signal for headers)
        2. Roman numerals at start (I., II., III., IV.)
        3. Ends with colon
        4. Too short (< 5 chars)
        5. All uppercase
        """
        import re
        
        if not cell_value:
            return False
        
        text = str(cell_value).strip()
        
        # Signal 1: Bold formatting (strongest signal)
        is_bold = False
        if hasattr(cell, 'font') and cell.font and hasattr(cell.font, 'bold'):
            is_bold = bool(cell.font.bold)
        
        # Signal 2: Contains Roman numerals at start (I., II., III., IV.)
        has_roman_numeral = bool(re.match(r'^[IVX]+\.', text))
        
        # Signal 3: Ends with colon (common for section headers)
        ends_with_colon = text.endswith(':')
        
        # Signal 4: Too short (< 5 chars) - likely just "I.", "II."
        is_too_short = len(text) < 5
        
        # Signal 5: All uppercase (REVENUE, EXPENSES)
        is_all_caps = text.isupper() and len(text) > 3
        
        # Decision logic:
        # If bold AND (has roman numeral OR ends with colon OR too short) → heading
        if is_bold and (has_roman_numeral or ends_with_colon or is_too_short):
            return True
        
        # If has roman numeral → likely heading
        if has_roman_numeral:
            return True
        
        # If ends with colon AND is bold → heading
        if ends_with_colon and is_bold:
            return True
        
        # If all caps AND bold → heading
        if is_all_caps and is_bold:
            return True
        
        # If too short AND bold → heading
        if is_too_short and is_bold:
            return True
        
        return False
    
    def _is_formula_cell(self, cell) -> bool:
        """
        Detect if a cell contains a formula.
        """
        # Method 1: Check data type
        if hasattr(cell, 'data_type') and cell.data_type == 'f':  # 'f' = formula
            return True
        
        # Method 2: Check if value starts with =
        if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
            return True
        
        return False
    
    def _read_template_periods(self, ws) -> List[str]:
        """
        Read period headers from template Row 2 (columns B onwards).
        
        Args:
            ws: Worksheet object
        
        Returns:
            List of period keys like ["30.06.2025", "31.03.2025_Y", "30.06.2024"]
        """
        import re
        
        periods = []
        
        try:
            # Read Row 2, starting from column B (index 2)
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=2, column=col_idx)
                if cell.value:
                    header_text = str(cell.value).strip()
                    print(f"header text: {header_text}") # Debug print
                    if header_text:
                        # Parse period from header
                        period_key = self._parse_period_from_header(header_text)
                        if period_key and period_key not in periods:
                            periods.append(period_key)
            
            _log.info(f"Read {len(periods)} periods from template Row 2: {periods}")
            return periods
            
        except Exception as e:
            _log.warning(f"Failed to read periods from template: {str(e)}")
            return []
    
    def _parse_period_from_header(self, header_text: str) -> str:
        """
        Convert any period header text to normalized key format.
        
        Examples:
        - "30.06.2025 Q" → "30_06_2025_q"
        - "31.03.2025 Y" → "31_03_2025_y"
        - "Quarter Ended June 30, 2025" → "quarter_ended_june_30_2025"
        - "Year Ended March 31, 2025" → "year_ended_march_31_2025"
        - "3M 30th June 2025" → "3m_30th_june_2025"
        - "12M FY 2025" → "12m_fy_2025"
        
        Args:
            header_text: Original period header from template
        
        Returns:
            Normalized key (lowercase, underscores, no special chars)
        """
        import re
        
        if not header_text:
            return None
        
        # Convert to lowercase
        normalized = str(header_text).strip().lower()
        
        # Replace dots with underscores (for dates like 30.06.2025)
        normalized = normalized.replace('.', '_')
        
        # Replace any whitespace with underscores
        normalized = re.sub(r'\s+', '_', normalized)
        
        # Remove special characters (except underscores)
        normalized = re.sub(r'[^a-z0-9_]', '', normalized)
        
        # Remove multiple consecutive underscores
        normalized = re.sub(r'_+', '_', normalized)
        
        # Remove leading/trailing underscores
        normalized = normalized.strip('_')
        
        return normalized if normalized else None
    
    def _read_template_metrics(self, template_excel_path: Path) -> List[str]:
        """
        Read metric names from Excel template (Column A, starting from row 3).
        Skips section headings using multiple detection signals.
        
        Args:
            template_excel_path: Path to Excel template file
        
        Returns:
            List of metric strings like "Sale of goods (key: sale_of_goods)"
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(template_excel_path, data_only=True)
            ws = wb.active
            
            metrics = []
            skipped_headings = []
            
            # Start from row 3 (rows 1-2 are usually headers)
            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)  # Column A
                if cell.value:
                    metric_name = str(cell.value).strip()
                    
                    # Skip empty or very short cells
                    if not metric_name or len(metric_name) < 3:
                        continue
                    
                    # Check if it's a section heading
                    if self._is_section_heading(metric_name, cell):
                        skipped_headings.append(metric_name)
                        _log.debug(f"Skipping section heading (row {row_idx}): '{metric_name}'")
                        continue
                    
                    # Check if it's a formula cell
                    is_formula = self._is_formula_cell(cell)
                    
                    # Infer key from metric name
                    key = self._infer_key_from_metric(metric_name)
                    
                    # Include metrics (including formulas - they exist in PDF as calculated values)
                    metric_str = f"{metric_name} (key: {key})"
                    if is_formula:
                        metric_str += " [formula]"
                    
                    metrics.append(metric_str)
                    _log.debug(f"Including metric (row {row_idx}): '{metric_name}' → key '{key}'{' [formula]' if is_formula else ''}")
            
            _log.info(f"Read {len(metrics)} metrics from template: {template_excel_path.name}")
            if skipped_headings:
                _log.info(f"Skipped {len(skipped_headings)} section headings: {skipped_headings}")
            return metrics
            
        except Exception as e:
            _log.warning(f"Failed to read template metrics: {str(e)}")
            return []
    
    def _read_template_structure(self, template_excel_path: Path) -> Dict:
        """
        Read complete template structure: periods AND metrics.
        
        Args:
            template_excel_path: Path to Excel template file
        
        Returns:
            Dict with 'periods' and 'metrics' lists
        """
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(template_excel_path, data_only=True)
            ws = wb.active
            
            # Read periods from Row 2
            periods = self._read_template_periods(ws)
            
            # Read metrics from Column A (using updated method that skips headings)
            # We need to read directly here instead of calling _read_template_metrics
            # because that method returns formatted strings
            metrics = []
            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)
                if cell.value:
                    metric_name = str(cell.value).strip()
                    if not metric_name or len(metric_name) < 3:
                        continue
                    if self._is_section_heading(metric_name, cell):
                        continue
                    key = self._infer_key_from_metric(metric_name)
                    metrics.append(f"{metric_name} (key: {key})")
            
            return {
                'periods': periods,
                'metrics': metrics
            }
            
        except Exception as e:
            _log.warning(f"Failed to read template structure: {str(e)}")
            return {'periods': [], 'metrics': []}
    
    def _infer_key_from_metric(self, metric_name: str) -> str:
        """
        Infer standardized key from metric display name.
        
        Args:
            metric_name: Display name like "Sale of Goods"
        
        Returns:
            Inferred key like "sale_of_goods"
        """
        # Convert to lowercase and replace spaces with underscores
        key = metric_name.lower().strip()
        
        # Remove common punctuation
        key = key.replace('(', '').replace(')', '').replace(',', '')
        key = key.replace('&', 'and').replace('-', '_')
        
        # Replace multiple spaces with single underscore
        key = '_'.join(key.split())
        
        # Remove leading/trailing underscores
        key = key.strip('_')
        
        return key
    
    def extract_from_html(self, html_path: Path, company_name: str, 
                         template_excel_path: Path = None) -> Dict:
        """
        Extract financial data from HTML table.
        
        Args:
            html_path: Path to HTML file containing financial table
            company_name: Company name for context
            template_excel_path: Optional Excel template for guided extraction
        
        Returns:
            Dictionary with extracted financial data
        """
        content = self._read_file_content(html_path)
        return self._extract_with_openai(content, company_name, "HTML", template_excel_path)
    
    def extract_from_markdown(self, md_path: Path, company_name: str,
                             template_excel_path: Path = None) -> Dict:
        """
        Extract financial data from Markdown table.
        
        Args:
            md_path: Path to Markdown file containing financial table
            company_name: Company name for context
            template_excel_path: Optional Excel template for guided extraction
        
        Returns:
            Dictionary with extracted financial data
        """
        content = self._read_file_content(md_path)
        return self._extract_with_openai(content, company_name, "Markdown", template_excel_path)
    
    def extract_from_output_dir(self, output_dir: Path, company_name: str, 
                                preferred_format: str = "html", 
                                template_excel_path: Path = None) -> Dict:
        """
        Extract financial data from output directory (tries multiple formats).
        
        Args:
            output_dir: Directory containing parsed output files
            company_name: Company name
            preferred_format: Preferred format to try first ("html" or "markdown")
            template_excel_path: Optional Excel template for guided extraction
        
        Returns:
            Dictionary with extracted financial data
        """
        # Try preferred format first
        if preferred_format == "html":
            html_files = list(output_dir.glob("*-table-*.html"))
            if html_files:
                _log.info(f"Using HTML file: {html_files[0]}")
                return self.extract_from_html(html_files[0], company_name, template_excel_path)
        
        # Try markdown as fallback
        md_files = list(output_dir.glob("*-table-*.md"))
        if md_files:
            _log.info(f"Using Markdown file: {md_files[0]}")
            return self.extract_from_markdown(md_files[0], company_name, template_excel_path)
        
        # Try HTML if markdown was preferred but not found
        if preferred_format == "markdown":
            html_files = list(output_dir.glob("*-table-*.html"))
            if html_files:
                _log.info(f"Falling back to HTML file: {html_files[0]}")
                return self.extract_from_html(html_files[0], company_name, template_excel_path)
        
        raise FileNotFoundError(f"No suitable table files found in {output_dir}")
    
    def _extract_with_openai(self, content: str, company_name: str, format_type: str,
                            template_excel_path: Path = None) -> Dict:
        """
        Extract financial data using OpenAI API.
        
        Args:
            content: HTML or Markdown content
            company_name: Company name
            format_type: "HTML" or "Markdown"
            template_excel_path: Optional Excel template for guided extraction
        
        Returns:
            Extracted financial data dictionary
        """
        try:
            # Truncate content if too long (to stay within token limits)
            max_content_length = 30000  # ~7500 tokens
            if len(content) > max_content_length:
                _log.warning(f"Content truncated from {len(content)} to {max_content_length} chars")
                content = content[:max_content_length]
            
            # Check if template is provided
            template_structure = {'periods': [], 'metrics': []}
            if template_excel_path and template_excel_path.exists():
                template_structure = self._read_template_structure(template_excel_path)
            
            # Choose appropriate prompts based on template availability
            if template_structure['metrics']:
                periods_text = ", ".join(template_structure['periods']) if template_structure['periods'] else "ALL periods from table"
                _log.info(f"Using TEMPLATE-GUIDED extraction with {len(template_structure['metrics'])} metrics and {len(template_structure['periods'])} periods")
                
                metrics_text = "\n".join([f"- {m}" for m in template_structure['metrics']])
                
                print(f"Using template with {metrics_text} metrics, {periods_text} periods")
                
                # Use template-aware system prompt with metrics
                system_prompt = self.SYSTEM_PROMPT_WITH_TEMPLATE.format(
                    template_metrics=metrics_text
                )

                user_prompt = self.USER_PROMPT_WITH_TEMPLATE.format(
                    format=format_type,
                    company_name=company_name,
                    period_mapping=periods_text,
                    content=content
                )
            
            else:
                _log.info("Using STANDARD extraction (no template)")
                system_prompt = self.SYSTEM_PROMPT
                user_prompt = self.USER_PROMPT_TEMPLATE.format(
                    format=format_type,
                    company_name=company_name,
                    content=content
                )
            
            _log.info(f"Sending request to OpenAI ({self.model})...")
            
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.1,  # Low temperature for consistent extraction
                response_format={"type": "json_object"}  # Ensure JSON response
            )
            
            # Extract JSON from response
            json_str = response.choices[0].message.content
            json_str = self._clean_json_response(json_str)
            
            _log.info(f"Received response from OpenAI (tokens used: {response.usage.total_tokens})")
            
            # Parse and validate JSON
            data = json.loads(json_str)
            # data =  {'company_name': 'BRITANNIA', 'financial_data': [{'particular': 'Sale of goods', 'key': 'sale_of_goods_/_income_from_operation', 'values': {'3m_30th_june_2025': '4,357.64', '3m_31st_mar_2025': '4,218.90', '3m_30th_june_2024': '3,967.38', '12_m_fy_2025': '16,859.22'}}, {'particular': 'Other operating revenues', 'key': 'other_operating_revenues', 'values': {'3m_30th_june_2025': '95.10', '3m_31st_mar_2025': '63.61', '3m_30th_june_2024': '127.06', '12_m_fy_2025': '436.70'}}, {'particular': 'Total Revenue', 'key': 'total_revenue', 'values': {'3m_30th_june_2025': '4,452.74', '3m_31st_mar_2025': '4,282.51', '3m_30th_june_2024': '4,094.44', '12_m_fy_2025': '17,295.92'}}, {'particular': 'Other Income', 'key': 'other_income', 'values': {'3m_30th_june_2025': '54.11', '3m_31st_mar_2025': '71.58', '3m_30th_june_2024': '67.29', '12_m_fy_2025': '250.68'}}, {'particular': 'Total Income', 'key': 'total_income', 'values': {'3m_30th_june_2025': '4,506.85', '3m_31st_mar_2025': '4,354.09', '3m_30th_june_2024': '4,161.73', '12_m_fy_2025': '17,546.60'}}, {'particular': 'Cost of materials consumed', 'key': 'cost_of_materials_consumed', 'values': {'3m_30th_june_2025': '2,250.47', '3m_31st_mar_2025': '2,126.03', '3m_30th_june_2024': '1,922.74', '12_m_fy_2025': '8,608.64'}}, {'particular': 'Purchases of stock-in-trade', 'key': 'purchases_of_stock_in_trade', 'values': {'3m_30th_june_2025': '506.38', '3m_31st_mar_2025': '507.33', '3m_30th_june_2024': '462.63', '12_m_fy_2025': '1,993.16'}}, {'particular': 'Changes in inventories of finished goods, work-in-progress and stock-in-trade', 'key': 'changes_in_inventories_of_finished_goods_work_in_porgress_and_stockin_trade', 'values': {'3m_30th_june_2025': '(20.14)', '3m_31st_mar_2025': '11.50', '3m_30th_june_2024': '1.67', '12_m_fy_2025': '(67.96)'}}, {'particular': 'Employee benefits expense', 'key': 'employee_benefits_expense', 'values': {'3m_30th_june_2025': '203.72', '3m_31st_mar_2025': '127.51', '3m_30th_june_2024': '164.71', '12_m_fy_2025': '554.70'}}, {'particular': 'Finance costs', 'key': 'finance_costs', 'values': {'3m_30th_june_2025': '25.68', '3m_31st_mar_2025': '30.10', '3m_30th_june_2024': '28.58', '12_m_fy_2025': '137.10'}}, {'particular': 'Depreciation and amortisation expense', 'key': 'depreciation_and_amortisation_expense', 'values': {'3m_30th_june_2025': '74.88', '3m_31st_mar_2025': '73.83', '3m_30th_june_2024': '66.71', '12_m_fy_2025': '284.67'}}, {'particular': 'Total expenses', 'key': 'total_expenses', 'values': {'3m_30th_june_2025': '3,832.65', '3m_31st_mar_2025': '3,607.74', '3m_30th_june_2024': '3,461.44', '12_m_fy_2025': '14,654.04'}}, {'particular': 'Profit before tax', 'key': 'profit_before_tax', 'values': {'3m_30th_june_2025': '674.20', '3m_31st_mar_2025': '746.35', '3m_30th_june_2024': '675.65', '12_m_fy_2025': '2,867.77'}}, {'particular': 'Current tax', 'key': 'current_tax', 'values': {'3m_30th_june_2025': '180.38', '3m_31st_mar_2025': '181.08', '3m_30th_june_2024': '176.86', '12_m_fy_2025': '730.63'}}, {'particular': 'Total tax', 'key': 'total_tax', 'values': {'3m_30th_june_2025': '175.93', '3m_31st_mar_2025': '189.25', '3m_30th_june_2024': '173.57', '12_m_fy_2025': '737.05'}}, {'particular': 'Profit for the year', 'key': 'profit_for_the_year', 'values': {'3m_30th_june_2025': '498.27', '3m_31st_mar_2025': '557.10', '3m_30th_june_2024': '502.08', '12_m_fy_2025': '2,130.72'}}, {'particular': 'EBITDA', 'key': 'ebitda', 'values': {'3m_30th_june_2025': '', '3m_31st_mar_2025': '746.35', '3m_30th_june_2024': '', '12_m_fy_2025': '2,892.56'}}, {'particular': 'EBITDA Margin', 'key': 'ebitda_margin', 'values': {'3m_30th_june_2025': '', '3m_31st_mar_2025': '', '3m_30th_june_2024': '', '12_m_fy_2025': ''}}]}
            
            # Validate structure
            if 'financial_data' not in data:
                raise ValueError("Response missing 'financial_data' key")
            
            # Add extraction metadata
            data['metadata'] = {
                'extraction_method': 'openai',
                'model': self.model,
                'tokens_used': 'N/A',
                'source_format': format_type.lower(),
                'template_guided': bool(template_structure['metrics']),
                'template_metrics_count': len(template_structure['metrics']),
                'template_periods_count': len(template_structure.get('period_mapping', {})),
                'template_periods': list(template_structure.get('period_mapping', {}).keys()) if template_structure.get('period_mapping') else None
            }
            
            _log.info(f"Successfully extracted {len(data['financial_data'])} financial items")
            return data
            
        except json.JSONDecodeError as e:
            _log.error(f"Failed to parse JSON response: {str(e)}")
            # _log.error(f"Raw response: {json_str[:500]}...")
            raise ValueError(f"Invalid JSON response from OpenAI: {str(e)}")
        except Exception as e:
            _log.error(f"Error during OpenAI extraction: {str(e)}", exc_info=True)
            raise


def validate_financial_data(data: Dict) -> bool:
    """
    Validate extracted financial data structure.
    
    Args:
        data: Extracted financial data dictionary
    
    Returns:
        True if valid, raises ValueError otherwise
    """
    if not isinstance(data, dict):
        raise ValueError("Data must be a dictionary")
    
    if 'company_name' not in data:
        raise ValueError("Missing 'company_name' key")
    
    if 'financial_data' not in data:
        raise ValueError("Missing 'financial_data' key")
    
    if not isinstance(data['financial_data'], list):
        raise ValueError("'financial_data' must be a list")
    
    # Validate each financial item
    for i, item in enumerate(data['financial_data']):
        if not isinstance(item, dict):
            raise ValueError(f"Item {i} is not a dictionary")
        
        required_keys = ['particular', 'key', 'values']
        for key in required_keys:
            if key not in item:
                raise ValueError(f"Item {i} missing required key: {key}")
        
        if not isinstance(item['values'], dict):
            raise ValueError(f"Item {i} 'values' must be a dictionary")
    
    return True
