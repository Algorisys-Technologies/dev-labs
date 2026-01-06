"""
Excel/CSV Generator for Financial Data
Converts JSON financial data to formatted Excel and CSV files
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
import uuid

_log = logging.getLogger(__name__)


class FinancialExcelGenerator:
    """Generate Excel and CSV files from financial JSON data.
    
    Supports two modes:
    1. Fixed format (47-row standard layout)
    2. JSON template-based (customizable rows and columns)
    """
    
    # Period mapping: JSON keys to column positions (B=1, C=2, etc.)
    PERIOD_MAPPING = {
        '30.06.2025': ('B', 1, 'Unaudited Q1', '3M-30th Jun 2025'),
        '31.03.2025_Y': ('C', 2, 'FY 2025', '12M'),
        '31.03.2025': ('D', 3, 'Q4', '3M-31st Mar 2025'),
        '31.12.2024': ('E', 4, 'Q3', '3M-31st Dec 2024'),
        '30.09.2024': ('F', 5, 'Q2', '3M-30th Sept 2024'),
        '30.06.2024': ('G', 6, 'Unaudited Q1 FY 2024', '3M-30th Jun 2024'),
        '31.03.2024_Y': ('H', 7, 'FY 2024', '12M'),
        '31.03.2024': ('I', 8, 'Q4 FY 2024', '3M-31st Mar 2024'),
        '31.12.2023': ('J', 9, 'Q3 FY 2024', '3M-31st Dec 2023'),
        '30.09.2023': ('K', 10, 'Q2 FY 2024', '3M-30th Sept 2023'),
        '30.06.2023': ('L', 11, 'Q1 FY 2024', '3M-30th Jun 2023'),
    }
    
    # Key aliases - support both naming conventions
    KEY_ALIASES = {
        'sale_of_products': 'sale_of_goods',
        'sale_of_goods': 'sale_of_goods',
    }
    
    def __init__(self):
        """Initialize the generator."""
        self.data_map = {}
        self.company_name = ""
    
    def _normalize_key(self, key: str) -> str:
        """Normalize key using aliases."""
        return self.KEY_ALIASES.get(key, key)
    
    def _parse_number(self, value: str) -> float:
        """Parse number from string, handling brackets for negatives."""
        if not value or value.strip() == '':
            return 0.0
        
        value = str(value).strip()
        
        # Handle brackets for negatives: (123) = -123
        if value.startswith('(') and value.endswith(')'):
            value = '-' + value[1:-1]
        
        # Remove commas
        value = value.replace(',', '')
        
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _format_number(self, value: float, decimal_places: int = 2) -> str:
        """Format number with Indian comma style and brackets for negatives."""
        if value == 0:
            return '-'
        
        # Handle negative numbers
        is_negative = value < 0
        abs_value = abs(value)
        
        # Format with decimal places
        formatted = f"{abs_value:,.{decimal_places}f}"
        
        # Convert to Indian numbering system (lakhs/crores)
        # For simplicity, keeping comma format but can be enhanced
        
        if is_negative:
            return f"({formatted})"
        return formatted
    
    def _build_data_map(self, financial_data: List[Dict]) -> None:
        """Build data map from financial_data array for easy lookup."""
        self.data_map = {}
        
        for item in financial_data:
            key = self._normalize_key(item.get('key', ''))
            # Support both 'values' and 'periods' keys for backward compatibility
            values = item.get('values', item.get('periods', {}))
            
            if key:
                self.data_map[key] = values
                _log.debug(f"Mapped key '{key}' with {len(values)} periods: {list(values.keys())}")
    
    def _get_value(self, key: str, period: str) -> float:
        """Get numeric value for a key and period."""
        normalized_key = self._normalize_key(key)
        
        if normalized_key in self.data_map:
            values = self.data_map[normalized_key]
            value = values.get(period, '')
            
            # If value is already a number, return it
            if isinstance(value, (int, float)):
                return float(value)
            
            # Otherwise parse it as a string
            return self._parse_number(value)
        
        return 0.0
    
    def _calculate_total(self, keys: List[str], period: str) -> float:
        """Calculate total from multiple keys."""
        total = 0.0
        for key in keys:
            total += self._get_value(key, period)
        return total
    
    def _create_excel_headers(self, ws) -> None:
        """Create Excel headers (rows 1-3)."""
        # Row 1: Company Name
        ws.merge_cells('B1:L1')
        cell = ws['B1']
        cell.value = self.company_name
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Row 2: Period Headers
        ws['A2'] = 'INR Crs'
        for period_key, (col_letter, _, header, _) in self.PERIOD_MAPPING.items():
            cell = ws[f'{col_letter}2']
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Row 3: Period Descriptions
        ws['A3'] = 'I. Revenue from operations'
        ws['A3'].font = Font(bold=True)
        ws['A3'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        for period_key, (col_letter, _, _, description) in self.PERIOD_MAPPING.items():
            cell = ws[f'{col_letter}3']
            cell.value = description
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    def _apply_cell_borders(self, ws, row: int, start_col: str = 'A', end_col: str = 'L') -> None:
        """Apply thin borders to cells in a row."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(ord(start_col), ord(end_col) + 1):
            cell = ws[f'{chr(col)}{row}']
            cell.border = thin_border
    
    def _apply_total_style(self, ws, row: int) -> None:
        """Apply total line styling (bold, double bottom border)."""
        for col in 'ABCDEFGHIJKL':
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='double')
            )
    
    def _set_row_data(self, ws, row: int, label: str, key: str, 
                      is_total: bool = False, is_section_header: bool = False) -> None:
        """Set data for a row across all periods."""
        # Set label
        ws[f'A{row}'] = label
        
        if is_section_header:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        elif is_total:
            ws[f'A{row}'].font = Font(bold=True)
        
        # Set data for each period
        for period_key, (col_letter, _, _, _) in self.PERIOD_MAPPING.items():
            value = self._get_value(key, period_key)
            cell = ws[f'{col_letter}{row}']
            
            if value != 0:
                cell.value = self._format_number(value)
            else:
                cell.value = '-'
            
            cell.alignment = Alignment(horizontal='right')
            
            if is_total:
                cell.font = Font(bold=True)
        
        # Apply borders
        self._apply_cell_borders(ws, row)
        
        if is_total:
            self._apply_total_style(ws, row)
    
    def generate_excel(self, json_data: Dict, output_path: Path, template_json_path: Path = None, template_excel_path: Path = None) -> bool:
        """
        Generate Excel file from JSON financial data.
        
        Args:
            json_data: Financial data in JSON format
            output_path: Path to save Excel file
            template_json_path: Optional path to JSON template file defining structure
            template_excel_path: Optional path to Excel template file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Extract data
            self.company_name = json_data.get('company_name', 'Financial Statement')
            financial_data = json_data.get('financial_data', [])
            
            # Build data map
            self._build_data_map(financial_data)
            
            # Check if Excel template is provided (takes priority)
            if template_excel_path and template_excel_path.exists():
                return self._generate_from_excel_template(json_data, output_path, template_excel_path)
            
            # Check if JSON template is provided
            if template_json_path and template_json_path.exists():
                return self._generate_from_json_template(json_data, output_path, template_json_path)
            
            # Otherwise use fixed format
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financial Statement"
            
            # Set column widths
            ws.column_dimensions['A'].width = 60
            for col in 'BCDEFGHIJKL':
                ws.column_dimensions[col].width = 15
            
            # Create headers (rows 1-3)
            self._create_excel_headers(ws)
            
            # Row 4-8: Revenue Section
            self._set_row_data(ws, 4, 'Sale of goods / Income from operations Domestic', 'sale_of_goods')
            self._set_row_data(ws, 5, 'Sale Exports', 'export_sales')
            self._set_row_data(ws, 6, 'Revenue from Services', 'service_revenue')
            self._set_row_data(ws, 7, 'Other operating revenues', 'other_operating_revenues')
            self._set_row_data(ws, 8, 'Total Revenue', 'revenue_from_operations', is_total=True)
            
            # Row 9: Other Income
            self._set_row_data(ws, 9, 'II. Other income', 'other_income')
            
            # Row 10-13: Total Income Section
            self._set_row_data(ws, 10, 'III. Total Income (I+II)', 'total_income')
            ws['A11'] = 'Sale of Goods Growth YOY'
            ws['A12'] = 'Total Revenue Growth YOY'
            ws['A13'] = 'Total Income Growth YOY'
            
            # Row 14: Empty
            
            # Row 15-32: Expenses Section
            ws['A15'] = 'IV. Expenses:'
            ws['A15'].font = Font(bold=True)
            ws['A15'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            self._set_row_data(ws, 16, 'Cost of materials consumed', 'cost_of_materials_consumed')
            self._set_row_data(ws, 17, 'Excise duty', 'excise_duty')
            self._set_row_data(ws, 18, 'Purchases of stock-in-trade', 'purchases_stock_in_trade')
            self._set_row_data(ws, 19, 'Changes in inventories of finished goods, work-in-progress and stock-in-trade', 
                              'changes_in_inventories')
            self._set_row_data(ws, 20, 'Employee benefits expense', 'employee_benefits_expense')
            self._set_row_data(ws, 21, 'Finance costs', 'finance_costs')
            self._set_row_data(ws, 22, 'Depreciation and amortisation expense', 'depreciation_amortisation_expense')
            self._set_row_data(ws, 23, 'Other expenses', 'other_expense')
            self._set_row_data(ws, 24, 'Advertising and promotion', 'advertising_expense')
            ws['A25'] = 'Others'
            self._set_row_data(ws, 26, 'Impairment', 'impairment_losses')
            ws['A27'] = 'Provision for contingencies'
            ws['A28'] = 'Corporate responsibilities'
            self._set_row_data(ws, 29, 'Total expenses', 'total_expenses', is_total=True)
            self._set_row_data(ws, 30, 'PBT before exp items', 'profit_before_exceptional_and_tax')
            # Row 31: Empty
            self._set_row_data(ws, 32, 'Exceptional items Gain/(Loss)', 'exceptional_item_expense')
            
            # Row 33: Empty
            
            # Row 34-35: Profit Before Tax
            self._set_row_data(ws, 34, 'V. Profit before tax (III-IV)', 'profit_before_tax', is_section_header=True)
            ws['A35'] = '%'
            
            # Row 36: Empty
            
            # Row 37-40: Tax Section
            ws['A37'] = 'VI. Tax expense:'
            ws['A37'].font = Font(bold=True)
            ws['A37'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            self._set_row_data(ws, 38, '(i) Current tax', 'current_tax')
            self._set_row_data(ws, 39, '(ii) Deferred tax/Income Tax of Prior years', 'deferred_tax')
            self._set_row_data(ws, 40, 'Total Tax', 'total_tax_expense', is_total=True)
            
            # Row 41: Net Profit
            self._set_row_data(ws, 41, 'VII. Profit for the year (V-VI)', 'net_profit', is_section_header=True)
            
            # Row 42: Empty
            
            # Row 43-44: EBITDA
            self._set_row_data(ws, 43, 'EBITDA', 'ebitda')
            ws['A44'] = 'EBITDA Margin'
            
            # Row 45: Empty
            
            # Row 46-47: Growth Metrics
            ws['A46'] = 'Volume Gr%'
            ws['A47'] = 'Price Gr%'
            
            # Apply borders to all used cells
            for row in range(1, 48):
                self._apply_cell_borders(ws, row)
            
            # Save workbook
            wb.save(output_path)
            _log.info(f"Excel file generated: {output_path}")
            return True
            
        except Exception as e:
            _log.error(f"Error generating Excel: {e}", exc_info=True)
            return False
    
    def _generate_from_excel_template(self, json_data: Dict, output_path: Path, template_excel_path: Path) -> bool:
        """
        Generate Excel by filling data in template using column mapping approach.
        
        Two modes supported:
        1. Placeholder mode: {{key[period]}} syntax
        2. Column mapping mode: Headers in Row 1, Metrics in Column A
        
        Args:
            json_data: Financial data with company_name and financial_data
            output_path: Path to save filled Excel file
            template_excel_path: Path to Excel template
        
        Returns:
            True if successful, False otherwise
        """
        try:
            import re
            from openpyxl import load_workbook
            from fuzzywuzzy import fuzz
            
            _log.info(f"Loading Excel template: {template_excel_path}")
            
            # Load template workbook
            wb = load_workbook(template_excel_path)
            
            # Build data map
            self.company_name = json_data.get('company_name', 'UNKNOWN')
            financial_data = json_data.get('financial_data', [])
            self._build_data_map(financial_data)
            
            # Try column mapping approach first, then fallback to placeholder mode
            placeholder_mode_used = False
            column_mapping_mode_used = False
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                _log.info(f"Processing sheet: {sheet_name}")
                
                # Try column mapping approach
                mapping_success = self._apply_column_mapping(ws)
                if mapping_success:
                    column_mapping_mode_used = True
                    _log.info(f"Applied column mapping to sheet: {sheet_name}")
                else:
                    # Fallback to placeholder mode
                    placeholder_success = self._apply_placeholder_mode(ws)
                    if placeholder_success:
                        placeholder_mode_used = True
                        _log.info(f"Applied placeholder mode to sheet: {sheet_name}")
            
            # Save filled workbook
            wb.save(output_path)
            
            if column_mapping_mode_used:
                _log.info(f"Excel generated using column mapping approach: {output_path}")
            elif placeholder_mode_used:
                _log.info(f"Excel generated using placeholder mode: {output_path}")
            else:
                _log.warning(f"Excel saved without data filling (no mappings or placeholders found): {output_path}")
            
            return True
            
        except Exception as e:
            _log.error(f"Error generating Excel from template: {str(e)}", exc_info=True)
            return False
    
    def _apply_column_mapping(self, ws) -> bool:
        """
        Apply column mapping approach to fill data.
        
        Detects:
        - Row 1: Company name in merged cells (B1:E1)
        - Row 2: Period headers (e.g., "30.06.2025 Q", "31.03.2025 Y")
        - Column A (from row 3): Metric names (e.g., "Sale of Goods", "Total Revenue")
        
        Returns:
            True if mapping was applied, False if not enough data to map
        """
        try:
            from fuzzywuzzy import fuzz
            import re
            
            # Step 0: Check for company name placeholder in row 1 (merged B1:E1)
            company_cell = ws['B1']
            if company_cell.value and isinstance(company_cell.value, str):
                if 'COMPANY_NAME_PLACEHOLDER' in company_cell.value or 'company_name' in company_cell.value.lower():
                    company_cell.value = self.company_name
                    _log.info(f"Filled company name in merged cell B1: {self.company_name}")
            
            # Step 1: Find header row (scan first 5 rows)
            header_row_num = None
            period_columns = {}  # {col_letter: period_key}
            
            _log.info("Scanning for period headers in first 5 rows...")
            
            for row_num in range(1, 6):
                row = list(ws[row_num])
                period_count = 0
                temp_period_columns = {}
                
                for col_idx, cell in enumerate(row[1:], start=2):  # Skip column A
                    if cell.value and isinstance(cell.value, str):
                        _log.debug(f"Checking cell {cell.column_letter}{row_num}: '{cell.value}'")
                        # Try to parse period from cell value
                        period_key = cell.value
                        if period_key:
                            col_letter = cell.column_letter
                            temp_period_columns[col_letter] = period_key
                            period_count += 1
                            _log.debug(f"  -> Parsed as period: {period_key}")
                
                # If we found 2+ periods, this is likely the header row
                if period_count >= 2:
                    header_row_num = row_num
                    period_columns = temp_period_columns
                    _log.info(f"Found header row at row {header_row_num} with {period_count} periods: {period_columns}")
                    break
            
            if not period_columns:
                _log.debug("No period columns detected - column mapping not applicable")
                return False
            
            # Step 2: Create metric name to key mapping
            metric_key_map = self._create_metric_key_map()
            
            # Step 3: Scan column A for metric names and fill data
            data_filled_count = 0
            metrics_start_row = header_row_num + 1
            
            _log.info(f"Scanning column A for metrics starting from row {metrics_start_row}...")
            _log.debug(f"Header row: {header_row_num}, Metrics start row: {metrics_start_row}")
            
            for row_num in range(metrics_start_row, ws.max_row + 1):
                metric_cell = ws[f'A{row_num}']
                
                if metric_cell.value and isinstance(metric_cell.value, str):
                    metric_name = str(metric_cell.value).strip()
                    
                    # Skip empty or very short strings
                    if len(metric_name) < 3:
                        continue
                    
                    _log.debug(f"Row {row_num}: Checking metric '{metric_name}'")
                    
                    # Find matching key for this metric
                    matched_key = self._match_metric_to_key(metric_name, metric_key_map)
                    
                    if matched_key:
                        _log.debug(f"  -> Matched to key: {matched_key}")
                        # Fill data for each period column
                        for col_letter, period_key in period_columns.items():
                            value = self._get_value(matched_key, period_key)
                            cell = ws[f'{col_letter}{row_num}']
                            
                            if value != 0:
                                formatted_value = self._format_number(value)
                                cell.value = formatted_value
                                data_filled_count += 1
                                _log.debug(f"  -> Filled {col_letter}{row_num} with {formatted_value} (key={matched_key}, period={period_key})")
                            else:
                                cell.value = '-'
                    else:
                        _log.debug(f"  -> No match found")
            
            if data_filled_count > 0:
                _log.info(f"Column mapping filled {data_filled_count} cells")
                return True
            else:
                _log.debug("Column mapping found structure but no data to fill")
                return False
                
        except Exception as e:
            _log.debug(f"Column mapping failed: {str(e)}")
            return False
    
    def _parse_period_from_header(self, header_text: str) -> str:
        """
        Parse period key from header text.
        
        Examples:
        - "30.06.2025 Q" → "30.06.2025"
        - "31.03.2025 Y" → "31.03.2025_Y"
        - "Q1 FY2026" → "30.06.2025" (if we can infer)
        - "FY 2025" → "31.03.2025_Y"
        """
        import re
        
        header_text = str(header_text).strip()
        
        # Pattern 1: Direct date format with Q or Y suffix
        # "30.06.2025 Q" or "31.03.2025 Y"
        match = re.search(r'(\d{2}\.\d{2}\.\d{4})\s*([QY])', header_text, re.IGNORECASE)
        if match:
            date_part = match.group(1)
            suffix = match.group(2).upper()
            if suffix == 'Y':
                return f"{date_part}_Y"
            return date_part
        
        # Pattern 2: Just the date without suffix (assume quarterly)
        # "30.06.2025"
        match = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', header_text)
        if match:
            return match.group(1)
        
        # Pattern 3: Quarterly labels
        # "Q1 FY2026", "Q1 2025", etc.
        match = re.search(r'Q[1-4].*?(20\d{2})', header_text, re.IGNORECASE)
        if match:
            year = match.group(1)
            # Map Q1 FY2026 → 30.06.2025
            if 'Q1' in header_text.upper():
                return f"30.06.{int(year)-1}"
            elif 'Q2' in header_text.upper():
                return f"30.09.{int(year)-1}"
            elif 'Q3' in header_text.upper():
                return f"31.12.{int(year)-1}"
            elif 'Q4' in header_text.upper():
                return f"31.03.{year}"
        
        # Pattern 4: Yearly labels
        # "FY 2025", "FY2025", "Year 2025"
        match = re.search(r'(?:FY|Year)\s*(20\d{2})', header_text, re.IGNORECASE)
        if match:
            year = match.group(1)
            return f"31.03.{year}_Y"
        
        return None
    
    def _create_metric_key_map(self) -> Dict[str, str]:
        """
        Create mapping from metric display names to keys.
        
        Returns:
            {display_name: key} dictionary
        """
        metric_map = {
            # Revenue section
            'sale of goods': 'sale_of_goods',
            'sale of products': 'sale_of_goods',
            'export sales': 'export_sales',
            'service revenue': 'service_revenue',
            'other operating revenue': 'other_operating_revenues',
            'other operating revenues': 'other_operating_revenues',
            'revenue from operations': 'revenue_from_operations',
            'total revenue from operations': 'revenue_from_operations',
            'total revenue': 'revenue_from_operations',
            'other income': 'other_income',
            'total income': 'total_income',
            
            # Expenses
            'cost of materials consumed': 'cost_of_materials_consumed',
            'cost of materials': 'cost_of_materials_consumed',
            'excise duty': 'excise_duty',
            'purchases of stock-in-trade': 'purchases_stock_in_trade',
            'purchases stock in trade': 'purchases_stock_in_trade',
            'changes in inventories': 'changes_in_inventories',
            'employee benefits expense': 'employee_benefits_expense',
            'employee benefit expense': 'employee_benefits_expense',
            'finance costs': 'finance_costs',
            'depreciation and amortisation': 'depreciation_amortisation_expense',
            'depreciation and amortization': 'depreciation_amortisation_expense',
            'other expenses': 'other_expense',
            'other expense': 'other_expense',
            'advertising expense': 'advertising_expense',
            'impairment losses': 'impairment_losses',
            'total expenses': 'total_expenses',
            
            # Profit & Tax
            'profit before exceptional items and tax': 'profit_before_exceptional_and_tax',
            'exceptional items': 'exceptional_item_expense',
            'profit before tax': 'profit_before_tax',
            'current tax': 'current_tax',
            'deferred tax': 'deferred_tax',
            'total tax expense': 'total_tax_expense',
            'tax expense': 'total_tax_expense',
            'net profit': 'net_profit',
            'profit for the period': 'net_profit',
            
            # OCI
            'other comprehensive income': 'other_comprehensive_income',
            'total comprehensive income': 'total_comprehensive_income',
            
            # Equity & EPS
            'paid-up equity share capital': 'paid_up_equity_share_capital',
            'equity share capital': 'paid_up_equity_share_capital',
            'other equity': 'other_equity',
            'eps basic': 'eps_basic',
            'eps (basic)': 'eps_basic',
            'basic eps': 'eps_basic',
            'eps diluted': 'eps_diluted',
            'eps (diluted)': 'eps_diluted',
            'diluted eps': 'eps_diluted',
        }
        
        return metric_map
    
    def _match_metric_to_key(self, metric_name: str, metric_map: Dict[str, str]) -> str:
        """
        Match metric name to key using fuzzy matching.
        
        Args:
            metric_name: Display name from Excel
            metric_map: Mapping dictionary
        
        Returns:
            Matched key or None
        """
        from fuzzywuzzy import fuzz
        
        metric_name_lower = metric_name.lower().strip()
        
        # Try exact match first
        if metric_name_lower in metric_map:
            return metric_map[metric_name_lower]
        
        # Try fuzzy matching
        best_match = None
        best_score = 0
        
        for display_name, key in metric_map.items():
            score = fuzz.ratio(metric_name_lower, display_name)
            if score > best_score and score > 80:  # 80% threshold
                best_score = score
                best_match = key
        
        if best_match:
            _log.debug(f"Matched '{metric_name}' to '{best_match}' (score: {best_score})")
        
        return best_match
    
    def _apply_placeholder_mode(self, ws) -> bool:
        """
        Apply placeholder replacement mode.
        
        Finds {{key[period]}} placeholders and replaces with data.
        
        Returns:
            True if placeholders were replaced, False otherwise
        """
        import re
        
        placeholder_pattern = re.compile(r'\{\{([^}\[]+)(?:\[([^]]+)\])?\}\}')
        total_placeholders_found = 0
        total_placeholders_replaced = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    matches = placeholder_pattern.findall(str(cell.value))
                    
                    if matches:
                        total_placeholders_found += len(matches)
                        cell_value = str(cell.value)
                        
                        for match in matches:
                            key = match[0].strip()
                            period = match[1].strip() if match[1] else None
                            
                            if key == 'company_name':
                                replacement = self.company_name
                                total_placeholders_replaced += 1
                            elif period:
                                value = self._get_value(key, period)
                                replacement = self._format_number(value) if value != 0 else '-'
                                total_placeholders_replaced += 1
                            else:
                                replacement = f"{{{{ERROR: {key} needs period}}}}"
                                _log.warning(f"Placeholder {{{{key}}}} found without period")
                            
                            if period:
                                placeholder = "{{" + key + "[" + period + "]}}"
                            else:
                                placeholder = "{{" + key + "}}"
                            cell_value = cell_value.replace(placeholder, str(replacement))
                        
                        # Update cell value
                        try:
                            if cell_value and cell_value != '-':
                                numeric_value = cell_value.replace(',', '')
                                if '(' in numeric_value and ')' in numeric_value:
                                    numeric_value = '-' + numeric_value.strip('()')
                                cell.value = float(numeric_value)
                            else:
                                cell.value = cell_value
                        except (ValueError, AttributeError):
                            cell.value = cell_value
        
        if total_placeholders_found > 0:
            _log.info(f"Replaced {total_placeholders_replaced} of {total_placeholders_found} placeholders")
            return True
        
        return False
    
    def _generate_from_json_template_old(self, json_data: Dict, output_path: Path, template_json_path: Path) -> bool:
        """
        Generate Excel from JSON template specification.
        
        Template format:
        {
          "columns": [{"period": "30.06.2025", "label": "Q1 FY2026"}, ...],
          "rows": [{"label": "Sale of goods", "key": "sale_of_goods", "type": "data"}, ...]
        }
        
        Args:
            json_data: Financial data with company_name and financial_data
            output_path: Path to save Excel file
            template_json_path: Path to JSON template file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            _log.info(f"Generating Excel from JSON template: {template_json_path}")
            
            # Load template
            with open(template_json_path, 'r', encoding='utf-8') as f:
                template = json.load(f)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = template.get('template_name', 'Financial Statement')
            
            # Get template configuration
            columns = template.get('columns', [])
            rows_spec = template.get('rows', [])
            layout = template.get('layout', {})
            formatting = template.get('formatting', {})
            
            # Title row
            title_row = layout.get('title_row', 1)
            ws.merge_cells(f'A{title_row}:' + get_column_letter(len(columns) + 1) + str(title_row))
            title_cell = ws[f'A{title_row}']
            title_cell.value = self.company_name
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            # Period header row
            period_row = layout.get('period_header_row', 2)
            ws[f'A{period_row}'] = 'INR Crs'
            ws[f'A{period_row}'].font = Font(bold=True)
            
            # Column headers (periods)
            for col_idx, col_spec in enumerate(columns, start=2):  # Start from column B
                col_letter = get_column_letter(col_idx)
                header_cell = ws[f'{col_letter}{period_row}']
                header_cell.value = col_spec.get('label', col_spec.get('period'))
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal='center')
                header_cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                
                # Set column width
                ws.column_dimensions[col_letter].width = 15
            
            # Set Column A width
            ws.column_dimensions['A'].width = 60
            
            # Data rows
            data_start_row = layout.get('data_start_row', 3)
            current_row = data_start_row
            
            for row_spec in rows_spec:
                row_type = row_spec.get('type', 'data')
                label = row_spec.get('label', '')
                key = row_spec.get('key')
                
                # Set label in column A
                label_cell = ws[f'A{current_row}']
                label_cell.value = label
                
                # Apply formatting based on type
                if row_type == 'section_header':
                    fmt = formatting.get('section_header', {})
                    if fmt.get('bold', False):
                        label_cell.font = Font(bold=True)
                    if 'background_color' in fmt:
                        label_cell.fill = PatternFill(start_color=fmt['background_color'], 
                                                      end_color=fmt['background_color'], fill_type='solid')
                elif row_type == 'total':
                    fmt = formatting.get('total', {})
                    if fmt.get('bold', True):
                        label_cell.font = Font(bold=True)
                elif row_type == 'metric':
                    fmt = formatting.get('metric', {})
                    if fmt.get('italic', False):
                        label_cell.font = Font(italic=True)
                elif row_type == 'blank':
                    current_row += 1
                    continue
                
                # Fill data values if key is provided
                if key and row_type in ['data', 'total']:
                    for col_idx, col_spec in enumerate(columns, start=2):
                        period = col_spec.get('period')
                        col_letter = get_column_letter(col_idx)
                        
                        # Get value from data map
                        value = self._get_value(key, period)
                        
                        data_cell = ws[f'{col_letter}{current_row}']
                        if value != 0:
                            data_cell.value = self._format_number(value)
                        else:
                            data_cell.value = '-'
                        
                        data_cell.alignment = Alignment(horizontal='right')
                        
                        # Apply total formatting
                        if row_type == 'total':
                            data_cell.font = Font(bold=True)
                            data_cell.border = Border(bottom=Side(style='double'))
                
                # Apply borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for col_idx in range(1, len(columns) + 2):
                    ws.cell(row=current_row, column=col_idx).border = thin_border
                
                current_row += 1
            
            # Save workbook
            wb.save(str(output_path))
            _log.info(f"Excel file generated from JSON template: {output_path}")
            return True
            
        except Exception as e:
            _log.error(f"Error generating Excel from JSON template: {str(e)}", exc_info=True)
            return False
    
    def generate_csv(self, json_data: Dict, output_path: Path) -> bool:
        """
        Generate CSV file from JSON financial data.
        
        Args:
            json_data: Financial data in JSON format
            output_path: Path to save CSV file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Extract data
            self.company_name = json_data.get('company_name', 'Financial Statement')
            financial_data = json_data.get('financial_data', [])
            
            # Build data map
            self._build_data_map(financial_data)
            
            # Prepare rows
            rows = []
            
            # Row 1: Company Name
            row1 = [''] + [self.company_name] + [''] * 10
            rows.append(row1)
            
            # Row 2: Period Headers
            row2 = ['INR Crs']
            for period_key, (_, _, header, _) in sorted(self.PERIOD_MAPPING.items(), 
                                                        key=lambda x: x[1][1]):
                row2.append(header)
            rows.append(row2)
            
            # Row 3: Period Descriptions
            row3 = ['I. Revenue from operations']
            for period_key, (_, _, _, description) in sorted(self.PERIOD_MAPPING.items(), 
                                                             key=lambda x: x[1][1]):
                row3.append(description)
            rows.append(row3)
            
            # Helper function to create data row
            def create_row(label: str, key: str) -> List[str]:
                row = [label]
                for period_key, (_, _, _, _) in sorted(self.PERIOD_MAPPING.items(), 
                                                       key=lambda x: x[1][1]):
                    value = self._get_value(key, period_key)
                    if value != 0:
                        row.append(self._format_number(value))
                    else:
                        row.append('-')
                return row
            
            # Add all rows according to structure
            rows.append(create_row('Sale of goods / Income from operations Domestic', 'sale_of_goods'))
            rows.append(create_row('Sale Exports', 'export_sales'))
            rows.append(create_row('Revenue from Services', 'service_revenue'))
            rows.append(create_row('Other operating revenues', 'other_operating_revenues'))
            rows.append(create_row('Total Revenue', 'revenue_from_operations'))
            
            rows.append(create_row('II. Other income', 'other_income'))
            
            rows.append(create_row('III. Total Income (I+II)', 'total_income'))
            rows.append(['Sale of Goods Growth YOY'] + ['-'] * 11)
            rows.append(['Total Revenue Growth YOY'] + ['-'] * 11)
            rows.append(['Total Income Growth YOY'] + ['-'] * 11)
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(['IV. Expenses:'] + ['-'] * 11)
            rows.append(create_row('Cost of materials consumed', 'cost_of_materials_consumed'))
            rows.append(create_row('Excise duty', 'excise_duty'))
            rows.append(create_row('Purchases of stock-in-trade', 'purchases_stock_in_trade'))
            rows.append(create_row('Changes in inventories of finished goods, work-in-progress and stock-in-trade', 
                                  'changes_in_inventories'))
            rows.append(create_row('Employee benefits expense', 'employee_benefits_expense'))
            rows.append(create_row('Finance costs', 'finance_costs'))
            rows.append(create_row('Depreciation and amortisation expense', 'depreciation_amortisation_expense'))
            rows.append(create_row('Other expenses', 'other_expense'))
            rows.append(create_row('Advertising and promotion', 'advertising_expense'))
            rows.append(['Others'] + ['-'] * 11)
            rows.append(create_row('Impairment', 'impairment_losses'))
            rows.append(['Provision for contingencies'] + ['-'] * 11)
            rows.append(['Corporate responsibilities'] + ['-'] * 11)
            rows.append(create_row('Total expenses', 'total_expenses'))
            rows.append(create_row('PBT before exp items', 'profit_before_exceptional_and_tax'))
            rows.append([''] * 12)  # Empty row
            rows.append(create_row('Exceptional items Gain/(Loss)', 'exceptional_item_expense'))
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(create_row('V. Profit before tax (III-IV)', 'profit_before_tax'))
            rows.append(['%'] + ['-'] * 11)
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(['VI. Tax expense:'] + ['-'] * 11)
            rows.append(create_row('(i) Current tax', 'current_tax'))
            rows.append(create_row('(ii) Deferred tax/Income Tax of Prior years', 'deferred_tax'))
            rows.append(create_row('Total Tax', 'total_tax_expense'))
            
            rows.append(create_row('VII. Profit for the year (V-VI)', 'net_profit'))
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(create_row('EBITDA', 'ebitda'))
            rows.append(['EBITDA Margin'] + ['-'] * 11)
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(['Volume Gr%'] + ['-'] * 11)
            rows.append(['Price Gr%'] + ['-'] * 11)
            
            # Write CSV
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            
            _log.info(f"CSV file generated: {output_path}")
            return True
            
        except Exception as e:
            _log.error(f"Error generating CSV: {e}", exc_info=True)
            return False
    
    def generate_excel_consolidated(self, companies_data: List[Dict], output_path: Path, template_excel_path: Path = None) -> bool:
        """
        Generate consolidated Excel from multiple companies' financial data.
        
        Layout:
        - Column A: Metric names (fixed)
        - Columns B-E: Company 1 periods
        - Column F: Blank separator
        - Columns G-J: Company 2 periods
        - Column K: Blank separator
        - And so on...
        
        Args:
            companies_data: List of financial data dicts, each with company_name and financial_data
            output_path: Path to save consolidated Excel
            template_excel_path: Optional Excel template for formatting
        
        Returns:
            True if successful, False otherwise
        """
        try:
            if not companies_data:
                _log.error("No companies data provided for consolidation")
                return False
            
            # Load template or create new workbook
            if template_excel_path and template_excel_path.exists():
                wb = openpyxl.load_workbook(template_excel_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Consolidated Financial Statement"
            
            # Set column A width
            ws.column_dimensions['A'].width = 60
            
            # Get standard metrics list (from first company or use default)
            standard_metrics = self._get_standard_metrics()
            
            # Extract periods from template Excel if provided, otherwise from data
            all_periods = []
            period_headers = []
            
            if template_excel_path and template_excel_path.exists():
                # Extract periods from Row 2 of template
                _log.info("Extracting periods from template Excel Row 2")
                for col_idx in range(2, ws.max_column + 1):  # Start from column B (index 2)
                    cell = ws.cell(row=2, column=col_idx)
                    if cell.value:
                        header_text = str(cell.value).strip()
                        if header_text:  # Only process non-empty headers
                            # Parse period from header
                            period_headers.append(header_text)
                            period_key = self._parse_period_from_header(header_text)
                            if period_key and period_key not in all_periods:
                                all_periods.append(period_key)
                            elif not period_key:
                                # If parsing fails, use the header text as-is
                                all_periods.append(header_text)
                _log.info(f"Extracted {len(all_periods)} periods from template: {all_periods}")
            
            # If no periods from template or template not provided, extract from data
            if not all_periods:
                _log.info("Extracting periods from financial data")
                for company_data in companies_data:
                    financial_data = company_data.get('financial_data', [])
                    for item in financial_data:
                        if 'periods' in item:
                            for period_key in item['periods'].keys():
                                if period_key not in all_periods:
                                    all_periods.append(period_key)
                                    period_headers.append(period_key)
                        elif 'values' in item:
                            for period_key in item['values'].keys():
                                if period_key not in all_periods:
                                    all_periods.append(period_key)
                                    period_headers.append(period_key)
            
            _log.info(f"Using periods from data: {all_periods}")
            periods_to_show = all_periods
            _log.info(f"Using periods for consolidation: {periods_to_show}")
            
            # Calculate column layout
            # Each company gets 4 columns + 1 blank separator
            current_col = 2  # Start at column B (index 2)
            
            # Row 1: Company names (merged across their period columns)
            row_idx = 1
            for company_idx, company_data in enumerate(companies_data):
                company_name = company_data.get('company_name', f'Company {company_idx + 1}')
                
                # Calculate column range for this company
                start_col = current_col
                end_col = current_col + len(periods_to_show) - 1
                
                # Merge cells for company name
                start_letter = get_column_letter(start_col)
                end_letter = get_column_letter(end_col)
                ws.merge_cells(f'{start_letter}{row_idx}:{end_letter}{row_idx}')
                
                # Set company name
                cell = ws[f'{start_letter}{row_idx}']
                cell.value = company_name
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(bold=True, size=14, color='FFFFFF')
                
                # Set column widths for this company's period columns
                for col_offset in range(len(periods_to_show)):
                    col_letter = get_column_letter(current_col + col_offset)
                    ws.column_dimensions[col_letter].width = 15
                
                # Move to next company's starting position (4 columns + 1 blank)
                current_col += len(periods_to_show) + 1
            
            # Row 2: Period headers
            row_idx = 2
            current_col = 2
            
            for company_idx, company_data in enumerate(companies_data):
                # Add period headers for this company
                for period_idx, period in enumerate(period_headers):
                    col_letter = get_column_letter(current_col + period_idx)
                    cell = ws[f'{col_letter}{row_idx}']
                    # Format period for display (e.g., "30.06.2025 Q" -> "Q1 FY26")
                    cell.value = self._format_period_header(period)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                
                # Leave blank column
                current_col += len(periods_to_show) + 1

            
            # Rows 3+: Metrics and data
            row_idx = 3
            
            for metric in standard_metrics:
                metric_name = metric['name']
                metric_key = metric['key']
                is_bold = metric.get('is_bold', False)
                
                # Column A: Metric name
                cell = ws[f'A{row_idx}']
                cell.value = metric_name
                if is_bold:
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Fill data for each company
                current_col = 2
                
                for company_idx, company_data in enumerate(companies_data):
                    # Build data map for this company
                    self._build_data_map(company_data.get('financial_data', []))
                    
                    # Fill period values for this company
                    for period_idx, period in enumerate(periods_to_show):
                        col_letter = get_column_letter(current_col + period_idx)
                        cell = ws[f'{col_letter}{row_idx}']
                        
                        # Get value from data map
                        value = self._get_value(metric_key, period)
                        
                        if value and value != 0:
                            # Handle numeric values
                            if isinstance(value, (int, float)):
                                cell.value = value
                                cell.number_format = '#,##0.00;(#,##0.00)'
                            else:
                                # Try to parse as number
                                try:
                                    num_val = self._parse_number(str(value))
                                    if num_val != 0:
                                        cell.value = num_val
                                        cell.number_format = '#,##0.00;(#,##0.00)'
                                    else:
                                        cell.value = '-'
                                except:
                                    cell.value = str(value)
                        else:
                            cell.value = '-'
                        
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Move to next company's columns
                    current_col += len(periods_to_show) + 1
                
                row_idx += 1
            
            # Save workbook
            wb.save(output_path)
            _log.info(f"Consolidated Excel file generated: {output_path}")
            return True
            
        except Exception as e:
            _log.error(f"Error generating consolidated Excel: {e}", exc_info=True)
            return False
    
    def _get_standard_metrics(self) -> List[Dict]:
        """Get standard list of financial metrics."""
        return [
            {'name': 'Sale of goods / Income from operations Domestic', 'key': 'sale_of_goods', 'is_bold': False},
            {'name': 'Sale Exports', 'key': 'export_sales', 'is_bold': False},
            {'name': 'Revenue from Services', 'key': 'service_revenue', 'is_bold': False},
            {'name': 'Other operating revenues', 'key': 'other_operating_revenues', 'is_bold': False},
            {'name': 'Total Revenue', 'key': 'revenue_from_operations', 'is_bold': True},
            {'name': 'Other income', 'key': 'other_income', 'is_bold': False},
            {'name': 'Total Income', 'key': 'total_income', 'is_bold': True},
            {'name': 'Cost of materials consumed', 'key': 'cost_of_materials', 'is_bold': False},
            {'name': 'Purchases of stock-in-trade', 'key': 'purchases_stock_in_trade', 'is_bold': False},
            {'name': 'Changes in inventories', 'key': 'changes_in_inventories', 'is_bold': False},
            {'name': 'Employee benefits expense', 'key': 'employee_benefits', 'is_bold': False},
            {'name': 'Depreciation and amortization', 'key': 'depreciation_amortization', 'is_bold': False},
            {'name': 'Other expenses', 'key': 'other_expenses', 'is_bold': False},
            {'name': 'Total Expenses', 'key': 'total_expenses', 'is_bold': True},
            {'name': 'Profit before exceptional items and tax', 'key': 'profit_before_exceptional_items', 'is_bold': False},
            {'name': 'Exceptional items', 'key': 'exceptional_items', 'is_bold': False},
            {'name': 'Profit before tax', 'key': 'profit_before_tax', 'is_bold': True},
            {'name': 'Tax expense', 'key': 'tax_expense', 'is_bold': False},
            {'name': 'Profit for the period', 'key': 'net_profit', 'is_bold': True},
            {'name': 'Other comprehensive income', 'key': 'other_comprehensive_income', 'is_bold': False},
            {'name': 'Total comprehensive income', 'key': 'total_comprehensive_income', 'is_bold': True},
            {'name': 'Basic EPS', 'key': 'basic_eps', 'is_bold': False},
            {'name': 'Diluted EPS', 'key': 'diluted_eps', 'is_bold': False},
            {'name': 'EBITDA', 'key': 'ebitda', 'is_bold': True},
        ]
    
    def _format_period_header(self, period: str) -> str:
        """Format period string for header display - returns original period as-is."""
        # Return the original period string without formatting
        # This preserves the full period format from the template (e.g., "30.06.2025 Q")
        return period


class FileManager:
    """Manage generated Excel/CSV files with metadata."""
    
    def __init__(self, storage_dir: Path):
        """
        Initialize file manager.
        
        Args:
            storage_dir: Directory to store files and metadata
        """
        self.storage_dir = storage_dir
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        self.metadata_file = storage_dir / 'metadata.json'
        self.metadata = self._load_metadata()
    
    def _load_metadata(self) -> Dict:
        """Load metadata from disk."""
        if self.metadata_file.exists():
            try:
                with open(self.metadata_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                _log.error(f"Error loading metadata: {e}")
                return {}
        return {}
    
    def _save_metadata(self) -> None:
        """Save metadata to disk."""
        try:
            with open(self.metadata_file, 'w', encoding='utf-8') as f:
                json.dump(self.metadata, f, indent=2)
        except Exception as e:
            _log.error(f"Error saving metadata: {e}")
    
    def save_file(self, file_path: Path, company_name: str, file_type: str) -> str:
        """
        Save file and create metadata entry.
        
        Args:
            file_path: Path to the generated file
            company_name: Company name
            file_type: 'excel' or 'csv'
        
        Returns:
            Unique file ID
        """
        file_id = str(uuid.uuid4())
        timestamp = datetime.now().isoformat()
        file_size = file_path.stat().st_size
        
        # Copy file to storage with unique name
        extension = '.xlsx' if file_type == 'excel' else '.csv'
        stored_path = self.storage_dir / f"{file_id}{extension}"
        
        import shutil
        shutil.copy2(file_path, stored_path)
        
        # Create metadata
        self.metadata[file_id] = {
            'file_id': file_id,
            'company_name': company_name,
            'file_type': file_type,
            'original_name': file_path.name,
            'stored_path': str(stored_path),
            'created_at': timestamp,
            'file_size': file_size,
            'download_count': 0,
            'stored_path': str(stored_path)
        }
        
        self._save_metadata()
        _log.info(f"File saved with ID: {file_id}")
        
        return file_id
    
    def get_file(self, file_id: str) -> Optional[Dict]:
        """
        Get file metadata and increment download count.
        
        Args:
            file_id: File ID
        
        Returns:
            File metadata or None if not found
        """
        if file_id in self.metadata:
            self.metadata[file_id]['download_count'] += 1
            self._save_metadata()
            return self.metadata[file_id]
        return None
    
    def list_files(self, company_name: Optional[str] = None) -> List[Dict]:
        """
        List all files with optional filtering.
        
        Args:
            company_name: Filter by company name (optional)
        
        Returns:
            List of file metadata
        """
        files = list(self.metadata.values())
        
        if company_name:
            files = [f for f in files if f['company_name'].upper() == company_name.upper()]
        
        # Sort by creation date (newest first)
        files.sort(key=lambda x: x['created_at'], reverse=True)
        
        return files
    
    def delete_file(self, file_id: str) -> bool:
        """
        Delete file and its metadata.
        
        Args:
            file_id: File ID
        
        Returns:
            True if successful, False otherwise
        """
        if file_id in self.metadata:
            try:
                # Delete physical file
                stored_path = Path(self.metadata[file_id]['stored_path'])
                if stored_path.exists():
                    stored_path.unlink()
                
                # Remove metadata
                del self.metadata[file_id]
                self._save_metadata()
                
                _log.info(f"File deleted: {file_id}")
                return True
            except Exception as e:
                _log.error(f"Error deleting file: {e}")
                return False
        
        return False
    
    def cleanup_old_files(self, days: int = 30) -> int:
        """
        Clean up files older than specified days.
        
        Args:
            days: Delete files older than this many days
        
        Returns:
            Number of files deleted
        """
        from datetime import timedelta
        
        cutoff_date = datetime.now() - timedelta(days=days)
        deleted_count = 0
        
        files_to_delete = []
        for file_id, metadata in self.metadata.items():
            created_at = datetime.fromisoformat(metadata['created_at'])
            if created_at < cutoff_date:
                files_to_delete.append(file_id)
        
        for file_id in files_to_delete:
            if self.delete_file(file_id):
                deleted_count += 1
        
        _log.info(f"Cleaned up {deleted_count} old files")
        return deleted_count
